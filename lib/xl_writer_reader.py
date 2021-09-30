from openpyxl import Workbook
from nptyping import NDArray
from openpyxl.worksheet.worksheet import Worksheet
from app.classes import MatrixBullet
from openpyxl.styles import PatternFill
from app import db
from datetime import datetime

def write_plan_to_xl(matrix: NDArray[MatrixBullet]):
  wb = Workbook()

  create_sheets(wb)
  write_sheets(matrix, wb)

  wb.save('Sorek-Plan-Optimized.xlsx')


def write_sheets(matrix: NDArray[MatrixBullet], wb: Workbook):
  write_holidays_sheet(matrix, wb)
  write_taoz_sheet(matrix, wb)
  write_price_sheet(matrix, wb)
  write_total_production_amount_sheet(matrix, wb)
  write_total_energy_consumption_sheet(matrix, wb)
  write_production_amount_sheet(matrix, wb)
  write_taoz_cost_sheet(matrix, wb)
  write_secondary_taoz_cost_sheet(matrix, wb)
  write_se_sheet(matrix, wb)
  write_num_of_pumps_sheet(matrix, wb)
  write_kwh_energy_limit_sheet(matrix, wb)
  write_shut_down_sheet(matrix, wb)
  write_production_price_sheet(matrix, wb)


def write_holidays_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  ws = wb['holidays']

  ws.cell(row=1, column=1).value = 'Holidays'
  ws.cell(row=1, column=2).value = 'Date'
  ws.cell(row=1, column=3).value = 'Day'

  holidays = db.holidays.find_one({}, {'_id': 0})

  i = 2
  for holiday in holidays:
    ws.cell(row=i, column=1).value = holiday
    ws.cell(row=i, column=2).value = holidays[holiday]['date']
    day = datetime.strptime(holidays[holiday]['date'], '%d/%m/%Y').date()
    ws.cell(row=i, column=3).value = day.strftime('%A')
    i += 1


def write_production_price_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  north_ws = wb['north_production_price']
  south_ws = wb['south_production_price']

  write_time(north_ws)
  write_time(south_ws)
  north_ws.cell(row=1, column=25).value = 'Daily Sum'
  south_ws.cell(row=1, column=25).value = 'Daily Sum'
  
  for i in range(1, len(matrix) + 1):
    north_sum = 0
    south_sum = 0
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      color: PatternFill = color_cell(bullet)
      north_ws.cell(row=i + 1, column=j + 1).value = bullet.north_facility.production_price
      north_ws.cell(row=i + 1, column=j + 1).fill = color
      south_ws.cell(row=i + 1, column=j + 1).value = bullet.south_facility.production_price
      south_ws.cell(row=i + 1, column=j + 1).fill = color
      north_sum += bullet.north_facility.production_price
      south_sum += bullet.south_facility.production_price
    north_ws.cell(row=i + 1, column=j + 2).value = north_sum
    south_ws.cell(row=i + 1, column=j + 2).value = south_sum
    north_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%d/%m/%Y')
    south_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%d/%m/%Y')
    north_ws.cell(row=i + 1, column=j + 4).value = bullet.date.strftime('%A')
    south_ws.cell(row=i + 1, column=j + 4).value = bullet.date.strftime('%A')


def write_shut_down_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  north_ws = wb['north_shut_down']
  south_ws = wb['south_shut_down']

  write_time(north_ws)
  write_time(south_ws)
  
  for i in range(1, len(matrix) + 1):
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      color: PatternFill = color_cell(bullet)
      north_ws.cell(row=i + 1, column=j + 1).value = bullet.north_facility.shutdown
      north_ws.cell(row=i + 1, column=j + 1).fill = color
      south_ws.cell(row=i + 1, column=j + 1).value = bullet.south_facility.shutdown
      south_ws.cell(row=i + 1, column=j + 1).fill = color
    north_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    south_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    north_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')
    south_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')


def write_kwh_energy_limit_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  north_ws = wb['north_kwh_energy_limit']
  south_ws = wb['south_kwh_energy_limit']

  write_time(north_ws)
  write_time(south_ws)
  
  for i in range(1, len(matrix) + 1):
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      color: PatternFill = color_cell(bullet)
      north_ws.cell(row=i + 1, column=j + 1).value = bullet.north_facility.kwh_energy_limit
      north_ws.cell(row=i + 1, column=j + 1).fill = color
      south_ws.cell(row=i + 1, column=j + 1).value = bullet.south_facility.kwh_energy_limit
      south_ws.cell(row=i + 1, column=j + 1).fill = color
    north_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    south_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    north_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')
    south_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')


def write_total_production_amount_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  ws = wb['total_production_amount']

  write_time(ws)
  ws.cell(row=1, column=25).value = 'Daily Sum'
  
  for i in range(1, len(matrix) + 1):
    row_sum = 0
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      color: PatternFill = color_cell(bullet)
      ws.cell(row=i + 1, column=j + 1).value = bullet.north_facility.production_amount + bullet.south_facility.production_amount
      ws.cell(row=i + 1, column=j + 1).fill = color
      row_sum += bullet.north_facility.production_amount + bullet.south_facility.production_amount
    ws.cell(row=i + 1, column=j + 2).value = row_sum
    ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%d/%m/%Y')
    ws.cell(row=i + 1, column=j + 4).value = bullet.date.strftime('%A')


def write_num_of_pumps_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  north_ws = wb['north_num_of_pumps']
  south_ws = wb['south_num_of_pumps']

  write_time(north_ws)
  write_time(south_ws)
  
  for i in range(1, len(matrix) + 1):
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      color: PatternFill = color_cell(bullet)
      north_ws.cell(row=i + 1, column=j + 1).value = bullet.north_facility.number_of_pumps
      north_ws.cell(row=i + 1, column=j + 1).fill = color
      south_ws.cell(row=i + 1, column=j + 1).value = bullet.south_facility.number_of_pumps
      south_ws.cell(row=i + 1, column=j + 1).fill = color
    north_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    south_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    north_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')
    south_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')


def write_se_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  north_ws = wb['north_se']
  south_ws = wb['south_se']

  write_time(north_ws)
  write_time(south_ws)
  
  for i in range(1, len(matrix) + 1):
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      color: PatternFill = color_cell(bullet)
      north_ws.cell(row=i + 1, column=j + 1).value = bullet.north_facility.se_per_hour
      north_ws.cell(row=i + 1, column=j + 1).fill = color
      south_ws.cell(row=i + 1, column=j + 1).value = bullet.south_facility.se_per_hour
      south_ws.cell(row=i + 1, column=j + 1).fill = color
    north_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    south_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    north_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')
    south_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')


def write_secondary_taoz_cost_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  north_ws = wb['north_secondary_taoz_cost']
  south_ws = wb['south_secondary_taoz_cost']

  write_time(north_ws)
  write_time(south_ws)
  
  for i in range(1, len(matrix) + 1):
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      color: PatternFill = color_cell(bullet)
      north_ws.cell(row=i + 1, column=j + 1).value = bullet.north_facility.secondary_taoz_cost
      north_ws.cell(row=i + 1, column=j + 1).fill = color
      south_ws.cell(row=i + 1, column=j + 1).value = bullet.south_facility.secondary_taoz_cost
      south_ws.cell(row=i + 1, column=j + 1).fill = color
    north_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    south_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    north_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')
    south_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')


def write_taoz_cost_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  north_ws = wb['north_taoz_cost']
  south_ws = wb['south_taoz_cost']

  write_time(north_ws)
  write_time(south_ws)
  
  for i in range(1, len(matrix) + 1):
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      color: PatternFill = color_cell(bullet)
      north_ws.cell(row=i + 1, column=j + 1).value = bullet.north_facility.taoz_cost
      north_ws.cell(row=i + 1, column=j + 1).fill = color
      south_ws.cell(row=i + 1, column=j + 1).value = bullet.south_facility.taoz_cost
      south_ws.cell(row=i + 1, column=j + 1).fill = color
    north_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    south_ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    north_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')
    south_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')


def write_production_amount_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  north_ws = wb['north_production_amount']
  south_ws = wb['south_production_amount']

  write_time(north_ws)
  write_time(south_ws)
  north_ws.cell(row=1, column=25).value = 'Daily Sum'
  south_ws.cell(row=1, column=25).value = 'Daily Sum'
  
  for i in range(1, len(matrix) + 1):
    north_sum = 0
    south_sum = 0
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      color: PatternFill = color_cell(bullet)
      north_ws.cell(row=i + 1, column=j + 1).value = bullet.north_facility.production_amount
      north_ws.cell(row=i + 1, column=j + 1).fill = color
      south_ws.cell(row=i + 1, column=j + 1).value = bullet.south_facility.production_amount
      south_ws.cell(row=i + 1, column=j + 1).fill = color
      north_sum += bullet.north_facility.production_amount
      south_sum += bullet.south_facility.production_amount
    north_ws.cell(row=i + 1, column=j + 2).value = north_sum
    south_ws.cell(row=i + 1, column=j + 2).value = south_sum
    north_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%d/%m/%Y')
    south_ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%d/%m/%Y')
    north_ws.cell(row=i + 1, column=j + 4).value = bullet.date.strftime('%A')
    south_ws.cell(row=i + 1, column=j + 4).value = bullet.date.strftime('%A')


def write_total_energy_consumption_sheet(matrix: NDArray[MatrixBullet], wb: Workbook) -> None:
  ws = wb['total_energy_consumption']

  write_time(ws)
  ws.cell(row=1, column=25).value = 'Daily Sum'
  
  for i in range(1, len(matrix) + 1):
    for j in range(len(matrix[i - 1])):
      row_sum = 0
      bullet: MatrixBullet = matrix[i - 1, j]
      energy_consumption = (bullet.north_facility.production_amount * bullet.north_facility.se_per_hour) + (bullet.south_facility.production_amount * bullet.south_facility.se_per_hour)
      ws.cell(row=i + 1, column=j + 1).value = energy_consumption
      ws.cell(row=i + 1, column=j + 1).fill = color_cell(bullet)
      row_sum += energy_consumption
    ws.cell(row=i + 1, column=j + 2).value = row_sum
    ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%d/%m/%Y')
    ws.cell(row=i + 1, column=j + 4).value = bullet.date.strftime('%A')


def write_price_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  ws = wb['price']

  write_time(ws)
  ws.cell(row=1, column=25).value = 'Daily Sum'
  
  for i in range(1, len(matrix) + 1):
    row_sum = 0
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      ws.cell(row=i + 1, column=j + 1).value = bullet.price
      ws.cell(row=i + 1, column=j + 1).fill = color_cell(bullet)
      row_sum += bullet.price
    ws.cell(row=i + 1, column=j + 2).value = row_sum
    ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%d/%m/%Y')
    ws.cell(row=i + 1, column=j + 4).value = bullet.date.strftime('%A')


def write_taoz_sheet(matrix: NDArray[MatrixBullet], wb: Workbook):
  ws = wb['taoz']

  write_time(ws)
  
  for i in range(1, len(matrix) + 1):
    for j in range(len(matrix[i - 1])):
      bullet: MatrixBullet = matrix[i - 1, j]
      ws.cell(row=i + 1, column=j + 1).value = bullet.taoz.name
      ws.cell(row=i + 1, column=j + 1).fill = color_cell(bullet)
    ws.cell(row=i + 1, column=j + 2).value = bullet.date.strftime('%d/%m/%Y')
    ws.cell(row=i + 1, column=j + 3).value = bullet.date.strftime('%A')


def color_cell(bullet: MatrixBullet) -> None:
  """
  checks the taoz type of the bullet and returns a color matching the taoz
  """
  redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
  greenFill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
  orangeFill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

  if bullet.taoz.name == 'SHEFEL':
    color_to_return = greenFill
  elif bullet.taoz.name == 'PISGA':
    color_to_return = redFill
  else:
    color_to_return = orangeFill

  return color_to_return


def create_sheets(wb: Workbook) -> None:
  """
  create the sheets for the xl workbook
  """
  ws = wb.active
  ws.title = 'taoz'

  wb.create_sheet('holidays')

  wb.create_sheet('price')
  wb.create_sheet('total_production_amount')
  wb.create_sheet('total_energy_consumption')

  wb.create_sheet('north_production_amount')
  wb.create_sheet('north_taoz_cost')
  wb.create_sheet('north_secondary_taoz_cost')
  wb.create_sheet('north_se')
  wb.create_sheet('north_num_of_pumps')
  wb.create_sheet('north_kwh_energy_limit')
  wb.create_sheet('north_shut_down')
  wb.create_sheet('north_production_price')

  wb.create_sheet('south_production_amount')
  wb.create_sheet('south_taoz_cost')
  wb.create_sheet('south_secondary_taoz_cost')
  wb.create_sheet('south_se')
  wb.create_sheet('south_num_of_pumps')
  wb.create_sheet('south_kwh_energy_limit')
  wb.create_sheet('south_shut_down')
  wb.create_sheet('south_production_price')


def write_time(ws: Worksheet) -> None:
  for i in range(24):
    ws.cell(row=1, column=i + 1).value = f'{i}:00'